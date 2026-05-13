import openpyxl
import os

file_path = os.path.join(os.path.dirname(__file__), '..', '2026長庚盃 排名、個人團體對抗表(2026新版).xlsx')
wb = openpyxl.load_workbook(file_path, data_only=True)

ind_sheet_name = next((n for n in wb.sheetnames if '16強' in n), None)
ws = wb[ind_sheet_name]

print(f"--- 個人對抗 Bronze Match Search (Row 35 to 60, Col 22 to 30) ---")
for r in range(34, 60):
    row_data = []
    for c in range(22, 30): 
        val = ws.cell(row=r+1, column=c+1).value
        val_str = str(val).replace('\n', ' ') if val is not None else ""
        if val_str:
            row_data.append(f"[{c}]:{val_str}")
    if row_data:
        print(f"Row {r}: " + " | ".join(row_data))
