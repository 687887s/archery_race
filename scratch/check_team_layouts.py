import openpyxl
import os

file_path = os.path.join(os.path.dirname(__file__), '..', '2026長庚盃 排名、個人團體對抗表(2026新版).xlsx')
wb = openpyxl.load_workbook(file_path, data_only=True)

for name in wb.sheetnames:
    if '團體對抗' in name:
        ws = wb[name]
        val = ws.cell(row=5, column=9).value # Row 4 (Index 3), Col 8 (Index 8 is I) -> 1/4 Left P1 Name
        print(f"{name}: 1/4 Left P1 Name = {val}")
        val2 = ws.cell(row=10, column=17).value # Row 9 (Index 9), Col 16 (Index 16 is Q) -> Gold Match P1 Name
        print(f"  Gold Match P1 Name = {val2}")
