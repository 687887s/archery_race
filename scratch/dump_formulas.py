import openpyxl
import os

file_path = os.path.join(os.path.dirname(__file__), '..', '2026長庚盃 排名、個人團體對抗表(2026新版).xlsx')
wb = openpyxl.load_workbook(file_path) # NO data_only=True

team_sheet_name = next((n for n in wb.sheetnames if '傳統30' in n and '團體' in n), None)
ws = wb[team_sheet_name]

print(f"--- 團隊賽 第一場 Formulas (Row 4 to 6) ---")
for r in range(3, 6):
    row_data = []
    for c in range(5, 11): # Col F to K
        val = ws.cell(row=r+1, column=c+1).value
        val_str = str(val).replace('\n', ' ') if val is not None else ""
        row_data.append(f"[{c}]:{val_str}")
    print(f"Row {r}: " + " | ".join(row_data))
