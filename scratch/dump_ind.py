import openpyxl
import os

file_path = os.path.join(os.path.dirname(__file__), '..', '2026長庚盃 排名、個人團體對抗表(2026新版).xlsx')
wb = openpyxl.load_workbook(file_path, data_only=True)

ind_sheet_name = next((n for n in wb.sheetnames if '傳統30' in n and '16強' in n), None)
ws = wb[ind_sheet_name]

print(f"--- 個人對抗 1/2 決賽與決賽 (Row 10 to 45, Col 18 to 35) ---")
for r in range(9, 42):
    row_data = []
    # Dump columns 18(S) to 33(AH)
    for c in range(18, 34):
        val = ws.cell(row=r+1, column=c+1).value
        if val:
            val_str = str(val).replace('\n', ' ')
            row_data.append(f"[{c}]:{val_str}")
    if row_data:
        print(f"Row {r}: " + " | ".join(row_data))
