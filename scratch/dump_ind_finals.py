import openpyxl
import os

file_path = os.path.join(os.path.dirname(__file__), '..', '2026長庚盃 排名、個人團體對抗表(2026新版).xlsx')
wb = openpyxl.load_workbook(file_path, data_only=True)

# Find first individual match sheet
ind_sheet_name = next((n for n in wb.sheetnames if '16強' in n), None)
ws = wb[ind_sheet_name]

print(f"--- 個人對抗 Final/Bronze (Row 16 to 36, Col 22 to 26) ---")
for r in range(13, 35):
    row_data = []
    for c in range(21, 26): 
        val = ws.cell(row=r+1, column=c+1).value
        val_str = str(val).replace('\n', ' ') if val is not None else ""
        if val_str:
            row_data.append(f"[{c}]:{val_str}")
    if row_data:
        print(f"Row {r}: " + " | ".join(row_data))
