import openpyxl
import os

file_path = os.path.join(os.path.dirname(__file__), '..', '2026長庚盃 排名、個人團體對抗表(2026新版).xlsx')
wb = openpyxl.load_workbook(file_path, data_only=True)

team_sheet_name = next((n for n in wb.sheetnames if '傳統30' in n and '團體' in n and '對抗' in n), None)
ws = wb[team_sheet_name]

print(f"--- 團隊對抗 1/2 Left Wing (Row 8 to 16, Col 10 to 14) ---")
for r in range(7, 16):
    row_data = []
    for c in range(10, 15): # Col K to O
        val = ws.cell(row=r+1, column=c+1).value
        val_str = str(val).replace('\n', ' ') if val is not None else ""
        if val_str:
            row_data.append(f"[{c}]:{val_str}")
    if row_data:
        print(f"Row {r}: " + " | ".join(row_data))
