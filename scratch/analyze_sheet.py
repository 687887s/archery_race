import openpyxl
import json
import os

file_path = os.path.join(os.path.dirname(__file__), '..', '2026長庚盃 排名、個人團體對抗表(2026新版).xlsx')

print(f"Loading workbook: {file_path}")
wb = openpyxl.load_workbook(file_path, data_only=True)

def get_val_at(ws, r, c):
    cell = ws.cell(row=r+1, column=c+1)
    val = cell.value
    if val is None:
        return ""
    return str(val).strip()

ind_matches = []
team_matches = []

# 1. 測試個人賽讀取邏輯
ind_sheet_name = next((n for n in wb.sheetnames if '傳統30' in n and ('16強' in n or '對抗' in n)), None)
if ind_sheet_name:
    print(f"Found individual sheet: {ind_sheet_name}")
    ws = wb[ind_sheet_name]
    for i in range(8):
        r_start = 8 + (i * 4)
        p1_unit = get_val_at(ws, r_start, 12)
        p1_name = get_val_at(ws, r_start, 13)
        p2_unit = get_val_at(ws, r_start + 2, 12)
        p2_name = get_val_at(ws, r_start + 2, 13)
        target = get_val_at(ws, r_start + 1, 14)
        
        if p1_name or p2_name or p1_unit or p2_unit:
            ind_matches.append({
                "matchId": f"M-1/8-{i+1}",
                "player1": p1_name or p1_unit or 'TBD',
                "player2": p2_name or p2_unit or 'TBD',
                "unit1": p1_unit,
                "unit2": p2_unit,
                "target": target
            })

# 2. 測試團體賽讀取邏輯
# FIND THE EXACT BRACKET SHEET! It must contain '對抗'
team_sheet_name = next((n for n in wb.sheetnames if '傳統30' in n and '團體' in n and '對抗' in n), None)
if team_sheet_name:
    print(f"Found team sheet: {team_sheet_name}")
    ws = wb[team_sheet_name]
    # 左翼 1/4 Round
    for i in range(2):
        r_start = 3 + (i * 8)
        p1_unit = get_val_at(ws, r_start, 7) # H is 7
        p1_name = get_val_at(ws, r_start, 8) # I is 8
        p2_unit = get_val_at(ws, r_start + 4, 7)
        p2_name = get_val_at(ws, r_start + 4, 8)
        target = get_val_at(ws, r_start + 2, 9) # J is 9? No, in my plan I used K(10) for target? 
        # Wait, the screenshot J is 0 (score).
        
        if p1_unit or p2_unit or p1_name or p2_name:
            team_matches.append({
                "matchId": f"MT-1/4-L{i+1}",
                "player1": p1_unit or 'TBD',
                "player2": p2_unit or 'TBD',
                "names1": p1_name.replace('\n', ' '),
                "names2": p2_name.replace('\n', ' '),
                "target": target
            })
    # 右翼 1/4 Round
    for i in range(2):
        r_start = 3 + (i * 8)
        p1_unit = get_val_at(ws, r_start, 31) # AF
        p1_name = get_val_at(ws, r_start, 30) # AE
        p2_unit = get_val_at(ws, r_start + 4, 31)
        p2_name = get_val_at(ws, r_start + 4, 30)
        target = get_val_at(ws, r_start + 2, 28) # AC?
        
        if p1_unit or p2_unit or p1_name or p2_name:
            team_matches.append({
                "matchId": f"MT-1/4-R{i+1}",
                "player1": p1_unit or 'TBD',
                "player2": p2_unit or 'TBD',
                "names1": p1_name.replace('\n', ' '),
                "names2": p2_name.replace('\n', ' '),
                "target": target
            })

print("\n=== 個人對抗賽測試結果 ===")
print(json.dumps(ind_matches, indent=2, ensure_ascii=False))

print("\n=== 團體對抗賽測試結果 ===")
print(json.dumps(team_matches, indent=2, ensure_ascii=False))
