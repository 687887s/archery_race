import openpyxl
import os

file_path = os.path.join(os.path.dirname(__file__), '..', '2026長庚盃 排名、個人團體對抗表(2026新版).xlsx')
wb = openpyxl.load_workbook(file_path, data_only=True)

def get_val_at(ws, r, c):
    val = ws.cell(row=r+1, column=c+1).value
    return str(val).strip().replace('\n', ' ') if val is not None else ""

# 1. Individual
ind_sheet_name = next((n for n in wb.sheetnames if '傳統30' in n and ('16強' in n or '對抗' in n)), None)
if ind_sheet_name:
    print(f"\n=== Individual Sheet: {ind_sheet_name} ===")
    ws = wb[ind_sheet_name]
    
    print("--- 1/8 Round ---")
    for i in range(8):
        rStart = 8 + (i * 4)
        p1Unit = get_val_at(ws, rStart, 12); p1Name = get_val_at(ws, rStart, 13)
        p2Unit = get_val_at(ws, rStart+2, 12); p2Name = get_val_at(ws, rStart+2, 13)
        print(f"Match {i+1}: P1=[{p1Unit}]/[{p1Name}], P2=[{p2Unit}]/[{p2Name}]")

    print("--- 1/4 Round ---")
    for i in range(4):
        rStart = 9 + (i * 8)
        p1Unit = get_val_at(ws, rStart, 18); p1Name = get_val_at(ws, rStart, 19)
        p2Unit = get_val_at(ws, rStart+4, 18); p2Name = get_val_at(ws, rStart+4, 19)
        print(f"Match {i+1}: P1=[{p1Unit}]/[{p1Name}], P2=[{p2Unit}]/[{p2Name}]")

    print("--- 1/2 Round ---")
    for i in range(2):
        rStart = 11 + (i * 16)
        p1Unit = get_val_at(ws, rStart, 23); p1Name = get_val_at(ws, rStart, 24)
        p2Unit = get_val_at(ws, rStart+8, 23); p2Name = get_val_at(ws, rStart+8, 24)
        print(f"Match {i+1}: P1=[{p1Unit}]/[{p1Name}], P2=[{p2Unit}]/[{p2Name}]")

    print("--- Final / Bronze ---")
    gP1Unit = get_val_at(ws, 14, 28); gP1Name = get_val_at(ws, 14, 29)
    gP2Unit = get_val_at(ws, 16, 28); gP2Name = get_val_at(ws, 16, 29)
    print(f"Gold: P1=[{gP1Unit}]/[{gP1Name}], P2=[{gP2Unit}]/[{gP2Name}]")
    
    bP1Unit = get_val_at(ws, 30, 28); bP1Name = get_val_at(ws, 30, 29)
    bP2Unit = get_val_at(ws, 32, 28); bP2Name = get_val_at(ws, 32, 29)
    print(f"Bronze: P1=[{bP1Unit}]/[{bP1Name}], P2=[{bP2Unit}]/[{bP2Name}]")


# 2. Team
team_sheet_name = next((n for n in wb.sheetnames if '傳統30' in n and '團體' in n and '對抗' in n), None)
if team_sheet_name:
    print(f"\n=== Team Sheet: {team_sheet_name} ===")
    ws = wb[team_sheet_name]
    
    print("--- 1/4 Left ---")
    for i in range(2):
        rStart = 3 + (i * 8)
        p1Unit = get_val_at(ws, rStart, 7); p1Name = get_val_at(ws, rStart, 8)
        p2Unit = get_val_at(ws, rStart+4, 7); p2Name = get_val_at(ws, rStart+4, 8)
        print(f"Match {i+1}: P1=[{p1Unit}]/[{p1Name}], P2=[{p2Unit}]/[{p2Name}]")

    print("--- 1/4 Right ---")
    for i in range(2):
        rStart = 3 + (i * 8)
        p1Unit = get_val_at(ws, rStart, 31); p1Name = get_val_at(ws, rStart, 30)
        p2Unit = get_val_at(ws, rStart+4, 31); p2Name = get_val_at(ws, rStart+4, 30)
        print(f"Match {i+1}: P1=[{p1Unit}]/[{p1Name}], P2=[{p2Unit}]/[{p2Name}]")

    print("--- 1/2 Left/Right ---")
    l12Unit1 = get_val_at(ws, 7, 11); l12Name1 = get_val_at(ws, 7, 12)
    l12Unit2 = get_val_at(ws, 15, 11); l12Name2 = get_val_at(ws, 15, 12)
    print(f"Left 1/2: P1=[{l12Unit1}]/[{l12Name1}], P2=[{l12Unit2}]/[{l12Name2}]")
    
    r12Unit1 = get_val_at(ws, 7, 27); r12Name1 = get_val_at(ws, 7, 26)
    r12Unit2 = get_val_at(ws, 15, 27); r12Name2 = get_val_at(ws, 15, 26)
    print(f"Right 1/2: P1=[{r12Unit1}]/[{r12Name1}], P2=[{r12Unit2}]/[{r12Name2}]")

    print("--- Final / Bronze ---")
    gUnit1 = get_val_at(ws, 8, 14); gName1 = get_val_at(ws, 8, 15)
    gUnit2 = get_val_at(ws, 8, 22); gName2 = get_val_at(ws, 8, 21)
    print(f"Gold: P1=[{gUnit1}]/[{gName1}], P2=[{gUnit2}]/[{gName2}]")

    bUnit1 = get_val_at(ws, 12, 14); bName1 = get_val_at(ws, 12, 15)
    bUnit2 = get_val_at(ws, 12, 22); bName2 = get_val_at(ws, 12, 21)
    print(f"Bronze: P1=[{bUnit1}]/[{bName1}], P2=[{bUnit2}]/[{bName2}]")
