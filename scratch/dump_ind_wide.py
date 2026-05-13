import openpyxl
import os

file_path = os.path.join(os.path.dirname(__file__), '..', '2026長庚盃 排名、個人團體對抗表(2026新版).xlsx')
wb = openpyxl.load_workbook(file_path, data_only=True)

ind_sheet_name = next((n for n in wb.sheetnames if '傳統30' in n and '16強' in n), None)
ws = wb[ind_sheet_name]

print(f"--- 個人對抗 1/4, 1/2, Finals (Row 10, 12, 14) ---")
for r in [9, 11, 13, 14, 15, 16, 29, 30, 31, 32]:
    row_data = []
    for c in range(16, 30): # Col Q to AD
        val = ws.cell(row=r+1, column=c+1).value
        val_str = str(val).replace('\n', ' ') if val is not None else ""
        row_data.append(f"[{c}]:{val_str}")
    print(f"Row {r}: " + " | ".join(row_data))
